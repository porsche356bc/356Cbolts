from openpyxl import load_workbook
from yattag import Doc, indent

# Load our Excel File
wb = load_workbook("Porsche_Nuts_and_Bolts.xlsx")
# Getting an object of active sheet 1
ws = wb.worksheets[1]

# Returning returns a triplet
doc, tag, text = Doc().tagtext()

xml_header = '<?xml version="1.0" encoding="UTF-8"?>'
xml_schema = '<xs:schema xmlns:xs="http://www.w3.org/2001/XMLSchema"></xs:schema>'

# Appends the String to document
doc.asis(xml_header)
doc.asis(xml_schema)


with tag('partslist'):
	for row in ws.iter_rows(min_row=2, max_row=7, min_col=1, max_col=11):
		row = [cell.value for cell in row]
		with tag("part") :
			with tag("diagram_name"): text(row[0])
			with tag("ref"): text(row[1])
			with tag("catalog") : text(row[2])
			with tag("group"): text(row[3])
			with tag("item"): text(row[4])
			with tag("description"): text(row[5])
			with tag("partno"): text(row[6])
			with tag("qty"): text(row[7])
			with tag("DIN"): text(row[8])
			with tag("size"): text(row[9])
			with tag("finish"): text(row[10])

# diagram_name	ref	catalog	group	item	description	part-no	qty	DIN	Size	Finish
#       0          1     2      3       4         5          6     7   8   9        10 

result = indent(
	doc.getvalue(),
	indentation='\t',
	indent_text=False
)

with open("partslist.xml", "w") as f:
	f.write(result)
